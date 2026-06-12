"""Тест ручного пути сборки (skip_search + ручные цены/ТСН), без сети и FastAPI.

Запуск:  python -m engine.test_build_manual
"""
from __future__ import annotations

from engine.pipeline import JobInputs, JobState, run_pipeline
from engine.schemas import BuildRequest


class _NoBrowser:
    def __enter__(self):
        raise AssertionError("browser не должен вызываться при skip_search и ручном ТСН")

    def __exit__(self, *a):
        return False


def _factory():
    return _NoBrowser()


REQ = BuildRequest(
    object_name="ТЕСТ · демонстрационный объект",
    skip_search=True,
    use_tsn=True,
    positions=[
        {  # обычная позиция с 3 ручными ценами; ТСН ниже КАЦ → остаётся
            "number": 1, "name": "Счетчик трехфазный", "unit": "шт", "qty": 10,
            "tsn_base_price": 200.0, "tsn_coefficient": 58.2,  # 11640 < 24033 → включить
            "offers": [
                {"price_with_vat": 22028, "org_name": "ООО «Альфа»", "inn": "7700000001",
                 "kpp": "770001001", "url": "https://shop1.example/x"},
                {"price_with_vat": 24033, "org_name": "ООО «Гамма»",
                 "inn": "7700000003", "kpp": "770001001", "url": "https://shop2.example/x"},
                {"price_with_vat": 22073.74, "org_name": "ООО «Бета»", "inn": "7700000002",
                 "kpp": "770001001", "url": "https://shop3.example/x"},
            ],
        },
        {  # ТСН выше КАЦ → исключить из КАЦ
            "number": 2, "name": "Кабель ППГнг 1х95", "unit": "м", "qty": 500,
            "tsn_base_price": 60.0, "tsn_coefficient": 58.2,  # 3492 > 2677 → исключить
            "offers": [
                {"price_with_vat": 2361.47, "org_name": "ООО «Дельта»", "inn": "7700000004"},
                {"price_with_vat": 2677.96, "org_name": "ООО «Дзета»", "inn": "7700000006"},
            ],
        },
        {  # уникальное из письма — 1 строка, без ТКП
            "number": 3, "name": "Щит распределительный (спецзаказ)", "is_unique": True,
            "offers": [
                {"price_with_vat": 1167443.13, "org_name": "ООО «Спецзавод»",
                 "inn": "7700000099", "kpp": "770001001", "from_letter": True},
            ],
        },
    ],
)


def main() -> None:
    results = [p.to_result() for p in REQ.positions]
    manual_tsn = {p.number: (p.tsn_base_price, p.tsn_coefficient)
                  for p in REQ.positions if p.tsn_base_price is not None}
    job = JobState(job_id="manualtest", object_name=REQ.object_name)
    inputs = JobInputs(object_name=REQ.object_name, spec_results=results,
                       use_tsn=True, skip_search=True, manual_tsn=manual_tsn)
    run_pipeline(job, inputs, _factory)

    print("status:", job.status)
    assert job.status == "done", job.error
    for line in job.log:
        print("  ·", line)
    print("outputs:", job.outputs)

    # проверки
    import openpyxl
    ws = openpyxl.load_workbook(job.outputs["kac"])["ЭМ (ЭЗС)"]
    nums = [ws.cell(r, 1).value for r in range(9, ws.max_row + 1) if ws.cell(r, 1).value]
    print("позиции в КАЦ:", nums)
    assert 2 not in nums, "позиция 2 должна быть исключена (ТСН>КАЦ)"
    assert 1 in nums and 3 in nums
    # макс. цена позиции 1 последней (по возрастанию)
    g = [ws.cell(r, 7).value for r in range(9, 12)]
    print("цены поз.1 (по возрастанию):", g)
    assert g == sorted(g) and g[-1] == 24033
    print("\nВСЕ ПРОВЕРКИ ПРОЙДЕНЫ ✓")


if __name__ == "__main__":
    main()
