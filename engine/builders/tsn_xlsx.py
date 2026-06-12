"""Сборщик файла сравнения с ТСН-2001 (Excel) + логика фильтрации КАЦ.

Для каждой позиции:
  расценка ТСН-2001 в текущих ценах = базовая расценка * коэффициент пересчёта (с mos.ru),
  % = КАЦ_макс / ТСН_текущая * 100,
  решение: если ТСН_текущая > КАЦ_макс → «исключить (ТСН>КАЦ)», иначе «включить».

Возвращает множество номеров позиций, которые нужно ИСКЛЮЧИТЬ из КАЦ.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from ..models import TsnRow

_thin = Side(style="thin")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
FONT = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

COLUMNS = [
    ("№ п.п.", 6),
    ("Наименование", 40),
    ("Ед. изм.", 9),
    ("Кол-во", 9),
    ("Базовая расценка ТСН-2001", 15),
    ("Коэффициент пересчёта", 13),
    ("Расценка ТСН в текущих ценах", 15),
    ("Макс. цена из КАЦ", 14),
    ("% (КАЦ / ТСН)", 12),
    ("Решение", 22),
]


def _compute(row: TsnRow) -> TsnRow:
    if row.tsn_base_price is not None and row.coefficient is not None:
        row.tsn_current_price = round(row.tsn_base_price * row.coefficient, 2)
    if row.tsn_current_price and row.kac_max_price:
        row.ratio_percent = round(row.kac_max_price / row.tsn_current_price * 100, 1)
        if row.tsn_current_price > row.kac_max_price:
            row.decision = "исключить (ТСН>КАЦ)"
        else:
            row.decision = "включить"
    elif row.tsn_current_price is None:
        row.decision = "нет расценки ТСН"
    else:
        row.decision = "нет цены КАЦ"
    return row


def build_tsn(
    rows: list[TsnRow],
    object_name: str,
    out_path: str | Path,
) -> tuple[Path, set[int]]:
    """Собирает файл ТСН-2001. Возвращает (путь, множество № позиций к исключению из КАЦ)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ТСН-2001"

    ws.merge_cells("A1:J1")
    ws["A1"] = "Сравнение с базой ТСН-2001"
    ws["A1"].font = Font(name="Arial", size=11, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:J2")
    ws["A2"] = object_name
    ws["A2"].font = FONT_BOLD
    ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

    hdr = 4
    for i, (title, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=hdr, column=i, value=title)
        c.font = FONT_BOLD
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[ws.cell(row=hdr, column=i).column_letter].width = width
    ws.row_dimensions[hdr].height = 45

    excluded: set[int] = set()
    r = hdr + 1
    for row in rows:
        _compute(row)
        if row.decision.startswith("исключить"):
            excluded.add(row.number)
        vals = [row.number, row.name, row.unit, row.qty, row.tsn_base_price,
                row.coefficient, row.tsn_current_price, row.kac_max_price,
                row.ratio_percent, row.decision]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = FONT
            c.border = BORDER
            c.alignment = LEFT if i == 2 else CENTER
            if i in (5, 7, 8):
                c.number_format = "#,##0.00"
            elif i == 9:
                c.number_format = "0.0"
        r += 1

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return Path(out_path), excluded
