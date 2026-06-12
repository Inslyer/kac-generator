"""Сборщик КАЦ (Excel) в формате эталонного образца.

Структура листа (воспроизводит эталон «КАЦ ЭМ и АК ЭЗС»):
- A1:Q1  «Конъюнктурный анализ»
- A2:Q2  наименование объекта
- A3:Q3  «(наименование объекта строительства)»
- строка 5 (+6) — заголовки 17 колонок A..Q
- строка 7 — номера граф: 1,2,3,4,5,6,7,8,13,14,15,16,17,18,19,20,21
- с строки 9 — блоки позиций по 3 строки (искомые) или 1 строка (из письма)

Для искомой позиции колонки A,B,C,D,E,F,I,J объединяются на 3 строки блока;
G,H,K,L,M,N,O,P,Q заполняются построчно. Цены идут по возрастанию (макс. — последняя).
H = G / divisor (НДС).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..config import settings
from ..models import PositionResult

# Заголовки колонок (строка 5) и номера граф (строка 7)
HEADERS = [
    "№ п.п.",
    "Код строительного ресурса",
    "Наименование строительного ресурса, затрат",
    "Полное наименование строительного ресурса, затрат в обосновывающем документе",
    "Ед. изм.",
    "Ед. изм. строительного ресурса, затрат в обосновывающем документе",
    "Текущая отпускная цена за ед. изм. в обосновывающем документе",
    "Текущая отпускная цена за ед. изм. без НДС в руб. в соответствии с обосновывающим документом",
    "Год",
    "Квартал",
    "Наименование производителя/поставщика",
    "КПП организации",
    "ИНН организации",
    "Гиперссылка на веб-сайт производителя/поставщика",
    "Населенный пункт расположения склада производителя/поставщика",
    "Статус организации (Производитель (1) / Поставщик (2)",
    "№ страницы прайс-листа в томе",
]
GRAPH_NUMBERS = [1, 2, 3, 4, 5, 6, 7, 8, 13, 14, 15, 16, 17, 18, 19, 20, 21]

COL_WIDTHS = {  # из эталона
    "A": 5.89, "B": 12.0, "C": 28.55, "D": 28.55, "E": 9.89, "F": 10.0,
    "G": 13.78, "H": 13.66, "I": 8.89, "J": 8.0, "K": 19.66, "L": 14.0,
    "M": 14.0, "N": 12.33, "O": 13.33, "P": 8.89, "Q": 10.0,
}
NCOLS = 17  # A..Q

_thin = Side(style="thin")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
FONT = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
FONT_TITLE = Font(name="Arial", size=11, bold=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_TOP = Alignment(horizontal="right", vertical="top", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)


def _write_sheet_header(ws: Worksheet, object_name: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NCOLS)
    ws["A1"] = "Конъюнктурный анализ"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = object_name
    ws["A2"].font = FONT_BOLD
    ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws["A3"] = "(наименование объекта строительства)"
    ws["A3"].font = Font(name="Arial", size=10, italic=True)
    ws["A3"].alignment = Alignment(horizontal="center")

    # строка 5 (+6) — заголовки, строка 7 — номера граф
    for i, (title, num) in enumerate(zip(HEADERS, GRAPH_NUMBERS), start=1):
        ws.merge_cells(start_row=5, start_column=i, end_row=6, end_column=i)
        h = ws.cell(row=5, column=i, value=title)
        h.font = FONT_BOLD
        h.alignment = WRAP_CENTER
        h.border = BORDER
        ws.cell(row=6, column=i).border = BORDER
        n = ws.cell(row=7, column=i, value=num)
        n.font = FONT
        n.alignment = Alignment(horizontal="center")
        n.border = BORDER
    ws.row_dimensions[5].height = 71.4
    ws.row_dimensions[6].height = 71.4
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w


def _merge_block(ws: Worksheet, col: int, r0: int, r1: int) -> None:
    if r1 > r0:
        ws.merge_cells(start_row=r0, start_column=col, end_row=r1, end_column=col)


def _write_position(ws: Worksheet, row: int, res: PositionResult) -> int:
    """Пишет блок позиции начиная с `row`. Возвращает следующий свободный ряд."""
    pos = res.position
    offers = res.offers
    n = max(1, len(offers))
    r0, r1 = row, row + n - 1

    # объединяемые на весь блок колонки
    block_cols = {1: pos.number, 2: "", 3: pos.name, 4: pos.full_name,
                  5: pos.unit, 6: pos.unit, 9: res.year, 10: res.quarter}
    for col, val in block_cols.items():
        _merge_block(ws, col, r0, r1)
        c = ws.cell(row=r0, column=col, value=val if val != "" else None)
        c.font = FONT
        if col == 1:
            c.alignment = CENTER_TOP
            c.number_format = "#,##0"
        elif col in (5, 6, 9, 10):
            c.alignment = CENTER_TOP
        else:
            c.alignment = LEFT_TOP

    # построчные колонки: G,H,K,L,M,N,O,P,Q
    for k, offer in enumerate(offers):
        r = r0 + k
        req = offer.requisites
        ws.cell(row=r, column=7, value=round(offer.price_with_vat, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=8,
                value=offer.price_without_vat(settings.vat_divisor)).number_format = "0.00"
        ws.cell(row=r, column=11, value=req.name)
        ws.cell(row=r, column=12, value=req.kpp or None)
        ws.cell(row=r, column=13, value=req.inn or None)
        if offer.url:
            cell = ws.cell(row=r, column=14, value=offer.url)
            cell.hyperlink = offer.url
        ws.cell(row=r, column=15, value=req.city or None)
        ws.cell(row=r, column=16, value=int(req.status))
        if res.tkp_page is not None and not offer.from_letter:
            ws.cell(row=r, column=17, value=res.tkp_page)

    if not offers:  # позиция без цен (на всякий случай) — пустой блок
        pass

    # стиль/границы по всему блоку
    for r in range(r0, r1 + 1):
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=r, column=col)
            c.border = BORDER
            if c.font is None or c.font.name != "Arial":
                c.font = FONT
            if col in (7, 8):
                c.alignment = RIGHT_TOP
            elif col in (12, 13, 16, 17):
                c.alignment = CENTER_TOP
            elif col == 14:
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 30
    return r1 + 1


def build_kac(
    results_by_discipline: dict[str, list[PositionResult]],
    object_name: str,
    out_path: str | Path,
) -> Path:
    """Собирает книгу КАЦ: по листу на дисциплину (ЭМ, АК)."""
    wb = Workbook()
    wb.remove(wb.active)
    for discipline, results in results_by_discipline.items():
        sheet_title = f"{discipline} (ЭЗС)"
        ws = wb.create_sheet(title=sheet_title[:31])
        _write_sheet_header(ws, object_name)
        row = 9
        for res in results:
            if res.excluded_by_tsn:
                continue
            row = _write_position(ws, row, res)
    out_path = Path(out_path)
    wb.save(out_path)
    return out_path
